from fastapi import FastAPI, APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
import os
import math
import logging
import uuid
import json
import base64
import asyncio
from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from concurrent.futures import ThreadPoolExecutor

from pydantic import BaseModel, Field, EmailStr
from passlib.context import CryptContext
from jose import jwt, JWTError

# Firebase Admin
import firebase_admin
from firebase_admin import credentials, firestore

# Cloudinary
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors as pdf_colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

_executor = ThreadPoolExecutor(max_workers=8)


def required_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value.strip().strip('"').strip("'")


def optional_env(name: str, default: str) -> str:
    return os.environ.get(name, default).strip().strip('"').strip("'")

# ---------------------------------------------------------------------------
# Firebase Firestore init
# ---------------------------------------------------------------------------
_fb_json_str = os.environ.get("FIREBASE_CREDENTIALS_JSON")
_fb_json_b64 = os.environ.get("FIREBASE_CREDENTIALS_B64")
_fb_cred_path = optional_env("FIREBASE_CREDENTIALS_PATH", "firebase_credentials.json")

if _fb_json_b64:
    cred_dict = json.loads(base64.b64decode(_fb_json_b64.strip().strip('"').strip("'")).decode("utf-8"))
    cred = credentials.Certificate(cred_dict)
elif _fb_json_str:
    cred_dict = json.loads(_fb_json_str.strip().strip("'"))
    cred = credentials.Certificate(cred_dict)
else:
    cred_path = Path(_fb_cred_path)
    if not cred_path.is_absolute():
        cred_path = ROOT_DIR / cred_path
    if not cred_path.exists():
        raise RuntimeError(f"Firebase credentials file not found: {cred_path}")
    cred = credentials.Certificate(str(cred_path))

if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
db = firestore.client()

# ---------------------------------------------------------------------------
# Cloudinary init
# ---------------------------------------------------------------------------
cloudinary.config(
    cloud_name=required_env("CLOUDINARY_CLOUD_NAME"),
    api_key=required_env("CLOUDINARY_API_KEY"),
    api_secret=required_env("CLOUDINARY_API_SECRET"),
    secure=True,
)

# ---------------------------------------------------------------------------
# JWT / App config
# ---------------------------------------------------------------------------
JWT_SECRET = required_env("JWT_SECRET")
JWT_ALGORITHM = optional_env("JWT_ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(optional_env("ACCESS_TOKEN_EXPIRE_MINUTES", "10080"))
CORS_ORIGINS = [
    origin.strip()
    for origin in optional_env("CORS_ORIGINS", "*").split(",")
    if origin.strip()
]

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

app = FastAPI(title="SoundOps Pro API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("soundops")

CATEGORIES = ["Speakers", "Amplifiers", "Mixers", "Microphones", "Cables", "Lighting", "Accessories"]
CATEGORY_LABELS = {
    "Speakers": "Bocinas",
    "Amplifiers": "Amplificadores",
    "Mixers": "Consolas",
    "Microphones": "Microfonos",
    "Cables": "Cables",
    "Lighting": "Iluminacion",
    "Accessories": "Accesorios",
}
CONDITION_LABELS = {
    "operational": "Operativo",
    "maintenance": "Mantenimiento",
    "broken": "Fuera de servicio",
}
STATUS_LABELS = {
    "scheduled": "Programado",
    "active": "Activo",
    "completed": "Finalizado",
}


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def new_id() -> str:
    return str(uuid.uuid4())


# ---------------------------------------------------------------------------
# Cloudinary helpers
# ---------------------------------------------------------------------------
def upload_image_to_cloudinary(image_data: str, folder: str = "soundops") -> str:
    if not image_data:
        return image_data
    if image_data.startswith("http://") or image_data.startswith("https://"):
        return image_data
    try:
        result = cloudinary.uploader.upload(
            image_data,
            folder=folder,
            resource_type="image",
        )
        return result["secure_url"]
    except Exception as e:
        logger.error(f"Cloudinary upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Error al subir imagen: {str(e)}")


# ---------------------------------------------------------------------------
# Firestore helpers
# ---------------------------------------------------------------------------
def fs_doc_to_dict(doc) -> dict:
    if not doc.exists:
        return None
    return doc.to_dict()


def fs_query_to_list(query) -> list:
    return [doc.to_dict() for doc in query.stream()]


def sort_created_desc(rows: list) -> list:
    fallback = datetime.min.replace(tzinfo=timezone.utc)
    return sorted(rows, key=lambda row: row.get("created_at") or fallback, reverse=True)


def fmt_dt(value) -> str:
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return str(value or "")


def fmt_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def event_totals(event: dict) -> tuple[int, int, int]:
    total = sum(int(a.get("quantity", 0)) for a in event.get("assignments", []))
    returned = sum(int(a.get("returned", 0)) for a in event.get("assignments", []))
    return total, returned, total - returned


def fetch_owner_report_data(owner_id: str) -> tuple[list, list]:
    equipment = fs_query_to_list(db.collection("equipment").where("owner_id", "==", owner_id))
    events = fs_query_to_list(db.collection("events").where("owner_id", "==", owner_id))
    return sort_created_desc(equipment), sort_created_desc(events)


def style_excel_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="FF9F0A")
    header_font = Font(bold=True, color="000000")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)
    ws.freeze_panes = "A2"


def build_excel_report(equipment: list, events: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append([
        "Nombre", "Marca", "Categoria", "Estado", "Cantidad total", "Disponible",
        "En eventos", "Notas", "Imagen", "Creado",
    ])
    for item in equipment:
        total = int(item.get("quantity", 0))
        available = int(item.get("quantity_available", 0))
        ws.append([
            fmt_text(item.get("name")),
            fmt_text(item.get("brand")),
            CATEGORY_LABELS.get(item.get("category"), item.get("category", "")),
            CONDITION_LABELS.get(item.get("condition"), item.get("condition", "")),
            total,
            available,
            total - available,
            fmt_text(item.get("notes")),
            fmt_text(item.get("image")),
            fmt_dt(item.get("created_at")),
        ])
    style_excel_sheet(ws)

    ev_ws = wb.create_sheet("Eventos")
    ev_ws.append([
        "Evento", "Lugar", "Fecha", "Estado", "Total asignado", "Devuelto",
        "Pendiente", "Notas", "Creado",
    ])
    for event in events:
        total, returned, pending = event_totals(event)
        ev_ws.append([
            fmt_text(event.get("name")),
            fmt_text(event.get("venue")),
            fmt_text(event.get("date")),
            STATUS_LABELS.get(event.get("status"), event.get("status", "")),
            total,
            returned,
            pending,
            fmt_text(event.get("notes")),
            fmt_dt(event.get("created_at")),
        ])
    style_excel_sheet(ev_ws)

    assign_ws = wb.create_sheet("Asignaciones")
    assign_ws.append(["Evento", "Estado evento", "Equipo", "Cantidad", "Devuelto", "Pendiente"])
    for event in events:
        for assignment in event.get("assignments", []):
            qty = int(assignment.get("quantity", 0))
            returned = int(assignment.get("returned", 0))
            assign_ws.append([
                fmt_text(event.get("name")),
                STATUS_LABELS.get(event.get("status"), event.get("status", "")),
                fmt_text(assignment.get("equipment_name")),
                qty,
                returned,
                qty - returned,
            ])
    style_excel_sheet(assign_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_pdf_report(equipment: list, events: list) -> BytesIO:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "SoundOpsTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        textColor=pdf_colors.HexColor("#0F1115"),
        fontSize=20,
        leading=24,
    )
    small_style = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8, leading=10)
    body_style = ParagraphStyle("Body", parent=styles["BodyText"], fontSize=9, leading=11)

    story = [
        Paragraph("Reporte SoundOps Pro", title_style),
        Paragraph(f"Generado: {fmt_dt(now_utc())}", small_style),
        Spacer(1, 12),
        Paragraph("Inventario", styles["Heading2"]),
    ]

    inventory_rows = [["Equipo", "Categoria", "Estado", "Total", "Disponible", "En eventos"]]
    for item in equipment:
        total = int(item.get("quantity", 0))
        available = int(item.get("quantity_available", 0))
        inventory_rows.append([
            Paragraph(fmt_text(item.get("name")), body_style),
            CATEGORY_LABELS.get(item.get("category"), item.get("category", "")),
            CONDITION_LABELS.get(item.get("condition"), item.get("condition", "")),
            total,
            available,
            total - available,
        ])
    if len(inventory_rows) == 1:
        inventory_rows.append(["Sin equipos", "", "", "", "", ""])
    story.append(Table(inventory_rows, repeatRows=1, colWidths=[230, 110, 110, 55, 70, 70]))
    story[-1].setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#FF9F0A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), pdf_colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, pdf_colors.HexColor("#C9CED6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    story.extend([PageBreak(), Paragraph("Eventos y asignaciones", styles["Heading2"])])
    event_rows = [["Evento", "Lugar/fecha", "Estado", "Asignado", "Devuelto", "Pendiente", "Equipos"]]
    for event in events:
        total, returned, pending = event_totals(event)
        gear = ", ".join(
            f"{a.get('equipment_name', '')} ({a.get('returned', 0)}/{a.get('quantity', 0)})"
            for a in event.get("assignments", [])
        )
        event_rows.append([
            Paragraph(fmt_text(event.get("name")), body_style),
            Paragraph(f"{fmt_text(event.get('venue'))}<br/>{fmt_text(event.get('date'))}", body_style),
            STATUS_LABELS.get(event.get("status"), event.get("status", "")),
            total,
            returned,
            pending,
            Paragraph(gear or "Sin equipos asignados", body_style),
        ])
    if len(event_rows) == 1:
        event_rows.append(["Sin eventos", "", "", "", "", "", ""])
    story.append(Table(event_rows, repeatRows=1, colWidths=[150, 135, 80, 55, 55, 60, 220]))
    story[-1].setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#FF9F0A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), pdf_colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, pdf_colors.HexColor("#C9CED6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    doc.build(story)
    output.seek(0)
    return output


# ✅ Helper para correr queries Firestore en paralelo desde async
async def fs_parallel(*queries):
    """Ejecuta múltiples queries Firestore en paralelo usando un thread pool."""
    loop = asyncio.get_event_loop()
    results = await asyncio.gather(
        *[loop.run_in_executor(_executor, lambda q=q: fs_query_to_list(q), ) for q in queries]
    )
    return results


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class RegisterIn(BaseModel):
    email: EmailStr
    name: str
    password: str


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    id: str
    email: EmailStr
    name: str
    created_at: datetime


class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut


class EquipmentIn(BaseModel):
    name: str
    brand: Optional[str] = ""
    category: str = "Accessories"
    serial: Optional[str] = ""
    condition: str = "operational"
    quantity: int = 1
    image: Optional[str] = None
    notes: Optional[str] = ""


class Equipment(EquipmentIn):
    id: str = Field(default_factory=new_id)
    owner_id: str
    quantity_available: int = 0
    created_at: datetime = Field(default_factory=now_utc)


class EventIn(BaseModel):
    name: str
    venue: Optional[str] = ""
    date: Optional[str] = ""
    notes: Optional[str] = ""


class Assignment(BaseModel):
    equipment_id: str
    equipment_name: str
    quantity: int
    returned: int = 0


class Event(EventIn):
    id: str = Field(default_factory=new_id)
    owner_id: str
    status: str = "scheduled"
    assignments: List[Assignment] = []
    created_at: datetime = Field(default_factory=now_utc)


class AssignIn(BaseModel):
    equipment_id: str
    quantity: int


class StatusIn(BaseModel):
    status: str


class SplIn(BaseModel):
    spl_ref: float
    dist_ref: float = 1.0
    dist_target: float


class ImpedanceIn(BaseModel):
    impedance: float
    count: int
    mode: str


class AmpPowerIn(BaseModel):
    speaker_power: float
    headroom_db: float = 3.0


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def create_token(sub: str) -> str:
    expire = now_utc() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode({"sub": sub, "exp": expire}, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    cred_err = HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        uid = payload.get("sub")
        if not uid:
            raise cred_err
    except JWTError:
        raise cred_err
    doc = db.collection("users").document(uid).get()
    user = fs_doc_to_dict(doc)
    if not user:
        raise cred_err
    return user


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@api_router.get("/health")
async def health():
    return {"status": "ok", "service": "SoundOps Pro API"}


@api_router.post("/auth/register", response_model=TokenOut)
async def register(body: RegisterIn):
    existing = db.collection("users").where("email", "==", body.email.lower()).limit(1).stream()
    if any(True for _ in existing):
        raise HTTPException(status_code=409, detail="Email already registered")
    user = {
        "id": new_id(),
        "email": body.email.lower(),
        "name": body.name,
        "password_hash": pwd_context.hash(body.password),
        "created_at": now_utc(),
    }
    db.collection("users").document(user["id"]).set(user)
    token = create_token(user["id"])
    return {"access_token": token, "user": UserOut(**user)}


@api_router.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn):
    docs = db.collection("users").where("email", "==", body.email.lower()).limit(1).stream()
    user = next((d.to_dict() for d in docs), None)
    if not user or not pwd_context.verify(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    token = create_token(user["id"])
    return {"access_token": token, "user": UserOut(**user)}


@api_router.get("/auth/me", response_model=UserOut)
async def me(user: dict = Depends(get_current_user)):
    return UserOut(**user)


# ---------------------------------------------------------------------------
# Equipment routes
# ---------------------------------------------------------------------------
@api_router.get("/equipment/categories")
async def equipment_categories(user: dict = Depends(get_current_user)):
    docs = db.collection("equipment").where("owner_id", "==", user["id"]).stream()
    counts = {}
    for doc in docs:
        cat = doc.to_dict().get("category", "Accessories")
        counts[cat] = counts.get(cat, 0) + 1
    total = sum(counts.values())
    return {"total": total, "categories": [{"name": c, "count": counts.get(c, 0)} for c in CATEGORIES]}


@api_router.get("/equipment", response_model=List[Equipment])
async def list_equipment(
    category: Optional[str] = None,
    q: Optional[str] = None,
    skip: int = 0,
    limit: int = 20,
    user: dict = Depends(get_current_user),
):
    query = db.collection("equipment").where("owner_id", "==", user["id"])
    docs = fs_query_to_list(query)
    if category and category != "All":
        docs = [d for d in docs if d.get("category") == category]
    if q:
        q_lower = q.lower()
        docs = [d for d in docs if q_lower in d.get("name", "").lower()]
    docs = sort_created_desc(docs)
    docs = docs[skip: skip + min(limit, 50)]
    return [Equipment(**d) for d in docs]


@api_router.post("/equipment", response_model=Equipment)
async def create_equipment(body: EquipmentIn, user: dict = Depends(get_current_user)):
    image_url = None
    if body.image:
        image_url = upload_image_to_cloudinary(body.image, folder="soundops/equipment")
    item = Equipment(**body.dict(), owner_id=user["id"], quantity_available=body.quantity)
    item.image = image_url
    db.collection("equipment").document(item.id).set(item.dict())
    return item


@api_router.get("/equipment/{item_id}", response_model=Equipment)
async def get_equipment(item_id: str, user: dict = Depends(get_current_user)):
    doc = db.collection("equipment").document(item_id).get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return Equipment(**data)


@api_router.put("/equipment/{item_id}", response_model=Equipment)
async def update_equipment(item_id: str, body: EquipmentIn, user: dict = Depends(get_current_user)):
    doc_ref = db.collection("equipment").document(item_id)
    doc = doc_ref.get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Equipment not found")
    deployed = data["quantity"] - data.get("quantity_available", data["quantity"])
    if body.quantity < deployed:
        raise HTTPException(status_code=400, detail=f"Cannot set quantity below {deployed} units currently deployed")
    new_available = max(0, body.quantity - deployed)
    updates = body.dict()
    if body.image and not body.image.startswith("http"):
        updates["image"] = upload_image_to_cloudinary(body.image, folder="soundops/equipment")
    updates["quantity_available"] = new_available
    doc_ref.update(updates)
    data.update(updates)
    return Equipment(**data)


@api_router.delete("/equipment/{item_id}")
async def delete_equipment(item_id: str, user: dict = Depends(get_current_user)):
    doc_ref = db.collection("equipment").document(item_id)
    doc = doc_ref.get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Equipment not found")
    doc_ref.delete()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Events
# ---------------------------------------------------------------------------
@api_router.get("/events", response_model=List[Event])
async def list_events(status: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = db.collection("events").where("owner_id", "==", user["id"])
    docs = fs_query_to_list(query)
    if status:
        docs = [d for d in docs if d.get("status") == status]
    docs = sort_created_desc(docs)
    return [Event(**d) for d in docs]


@api_router.post("/events", response_model=Event)
async def create_event(body: EventIn, user: dict = Depends(get_current_user)):
    event = Event(**body.dict(), owner_id=user["id"])
    db.collection("events").document(event.id).set(event.dict())
    return event


@api_router.get("/events/{event_id}", response_model=Event)
async def get_event(event_id: str, user: dict = Depends(get_current_user)):
    doc = db.collection("events").document(event_id).get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Event not found")
    return Event(**data)


@api_router.patch("/events/{event_id}/status", response_model=Event)
async def set_event_status(event_id: str, body: StatusIn, user: dict = Depends(get_current_user)):
    doc_ref = db.collection("events").document(event_id)
    doc = doc_ref.get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Event not found")
    doc_ref.update({"status": body.status})
    data["status"] = body.status
    return Event(**data)


@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, user: dict = Depends(get_current_user)):
    doc_ref = db.collection("events").document(event_id)
    doc = doc_ref.get()
    data = fs_doc_to_dict(doc)
    if not data or data.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
    if data.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Solo puedes eliminar eventos finalizados")
    _, _, pending = event_totals(data)
    if pending > 0:
        raise HTTPException(status_code=400, detail="Devuelve todos los equipos antes de eliminar el evento")
    doc_ref.delete()
    return {"ok": True}


@api_router.post("/events/{event_id}/assign", response_model=Event)
async def assign_gear(event_id: str, body: AssignIn, user: dict = Depends(get_current_user)):
    event_ref = db.collection("events").document(event_id)
    event_doc = event_ref.get()
    event = fs_doc_to_dict(event_doc)
    if not event or event.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Event not found")
    eq_doc = db.collection("equipment").document(body.equipment_id).get()
    eq = fs_doc_to_dict(eq_doc)
    if not eq or eq.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Equipment not found")
    if body.quantity <= 0 or body.quantity > eq.get("quantity_available", 0):
        raise HTTPException(status_code=400, detail="Not enough units available")
    db.collection("equipment").document(body.equipment_id).update(
        {"quantity_available": firestore.Increment(-body.quantity)}
    )
    assignments = event.get("assignments", [])
    found = next((a for a in assignments if a["equipment_id"] == body.equipment_id), None)
    if found:
        found["quantity"] += body.quantity
    else:
        assignments.append({
            "equipment_id": body.equipment_id,
            "equipment_name": eq["name"],
            "quantity": body.quantity,
            "returned": 0,
        })
    event_ref.update({"assignments": assignments})
    movement_id = new_id()
    db.collection("movements").document(movement_id).set({
        "id": movement_id,
        "owner_id": user["id"],
        "type": "out",
        "equipment_id": body.equipment_id,
        "equipment_name": eq["name"],
        "event_id": event_id,
        "event_name": event["name"],
        "quantity": body.quantity,
        "created_at": now_utc(),
    })
    event["assignments"] = assignments
    return Event(**event)


@api_router.post("/events/{event_id}/return", response_model=Event)
async def return_gear(event_id: str, body: AssignIn, user: dict = Depends(get_current_user)):
    event_ref = db.collection("events").document(event_id)
    event_doc = event_ref.get()
    event = fs_doc_to_dict(event_doc)
    if not event or event.get("owner_id") != user["id"]:
        raise HTTPException(status_code=404, detail="Event not found")
    assignments = event.get("assignments", [])
    found = next((a for a in assignments if a["equipment_id"] == body.equipment_id), None)
    if not found:
        raise HTTPException(status_code=404, detail="Assignment not found")
    outstanding = found["quantity"] - found.get("returned", 0)
    if body.quantity <= 0 or body.quantity > outstanding:
        raise HTTPException(status_code=400, detail="Invalid return quantity")
    found["returned"] = found.get("returned", 0) + body.quantity
    db.collection("equipment").document(body.equipment_id).update(
        {"quantity_available": firestore.Increment(body.quantity)}
    )
    event_ref.update({"assignments": assignments})
    movement_id = new_id()
    db.collection("movements").document(movement_id).set({
        "id": movement_id,
        "owner_id": user["id"],
        "type": "in",
        "equipment_id": body.equipment_id,
        "equipment_name": found["equipment_name"],
        "event_id": event_id,
        "event_name": event["name"],
        "quantity": body.quantity,
        "created_at": now_utc(),
    })
    event["assignments"] = assignments
    return Event(**event)


# ---------------------------------------------------------------------------
# Dashboard — ✅ FIX: queries en paralelo (era ~5x más lento en serie)
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------
@api_router.get("/exports/report.xlsx")
async def export_excel(user: dict = Depends(get_current_user)):
    equipment, events = fetch_owner_report_data(user["id"])
    output = build_excel_report(equipment, events)
    headers = {"Content-Disposition": 'attachment; filename="soundops-reporte.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@api_router.get("/exports/report.pdf")
async def export_pdf(user: dict = Depends(get_current_user)):
    equipment, events = fetch_owner_report_data(user["id"])
    output = build_pdf_report(equipment, events)
    headers = {"Content-Disposition": 'attachment; filename="soundops-reporte.pdf"'}
    return StreamingResponse(output, media_type="application/pdf", headers=headers)


@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    uid = user["id"]
    loop = asyncio.get_event_loop()

    # Lanzar las 3 queries principales en paralelo
    def _get_equipment():
        return fs_query_to_list(db.collection("equipment").where("owner_id", "==", uid))

    def _get_events():
        return fs_query_to_list(db.collection("events").where("owner_id", "==", uid))

    def _get_movements():
        rows = fs_query_to_list(db.collection("movements").where("owner_id", "==", uid))
        return sort_created_desc(rows)[:8]

    equipment, all_events, recent = await asyncio.gather(
        loop.run_in_executor(_executor, _get_equipment),
        loop.run_in_executor(_executor, _get_events),
        loop.run_in_executor(_executor, _get_movements),
    )

    total_gear = len(equipment)
    total_units = sum(e.get("quantity", 0) for e in equipment)
    available_units = sum(e.get("quantity_available", 0) for e in equipment)
    deployed_units = total_units - available_units
    stock_alerts = [
        {
            "id": e["id"],
            "name": e["name"],
            "reason": "Out of stock" if e.get("quantity_available", 0) == 0 else "Needs service",
        }
        for e in equipment
        if e.get("quantity_available", 0) == 0 or e.get("condition") != "operational"
    ]
    active_events = sum(1 for e in all_events if e.get("status") != "completed")

    return {
        "total_gear": total_gear,
        "total_units": total_units,
        "deployed_units": deployed_units,
        "available_units": available_units,
        "active_events": active_events,
        "stock_alerts": stock_alerts[:10],
        "recent_movements": recent,
    }


# ---------------------------------------------------------------------------
# Acoustic Calculator
# ---------------------------------------------------------------------------
@api_router.post("/calc/spl")
async def calc_spl(body: SplIn, user: dict = Depends(get_current_user)):
    if body.dist_ref <= 0 or body.dist_target <= 0:
        raise HTTPException(status_code=400, detail="Distances must be greater than zero")
    loss = 20 * math.log10(body.dist_target / body.dist_ref)
    spl_target = body.spl_ref - loss
    return {"spl_target": round(spl_target, 1), "loss_db": round(loss, 1), "note": "Free-field inverse-square law (−6 dB per distance doubling)."}


@api_router.post("/calc/impedance")
async def calc_impedance(body: ImpedanceIn, user: dict = Depends(get_current_user)):
    if body.count < 1 or body.impedance <= 0:
        raise HTTPException(status_code=400, detail="Invalid speaker configuration")
    if body.mode == "series":
        total = body.impedance * body.count
    elif body.mode == "parallel":
        total = body.impedance / body.count
    else:
        raise HTTPException(status_code=400, detail="Mode must be series or parallel")
    warning = "Total load below 4Ω — verify amplifier minimum impedance to avoid damage." if total < 4 else None
    return {"total_impedance": round(total, 2), "mode": body.mode, "warning": warning}


@api_router.post("/calc/amp-power")
async def calc_amp_power(body: AmpPowerIn, user: dict = Depends(get_current_user)):
    if body.speaker_power <= 0:
        raise HTTPException(status_code=400, detail="Speaker power must be greater than zero")
    factor = 10 ** (body.headroom_db / 10)
    recommended = body.speaker_power * factor
    return {
        "recommended_power": round(recommended, 1),
        "headroom_db": body.headroom_db,
        "range_min": round(body.speaker_power, 1),
        "range_max": round(recommended, 1),
        "note": f"Allow ~{body.headroom_db} dB headroom above the speaker's continuous rating.",
    }


# ---------------------------------------------------------------------------
# ✅ NUEVO: Ping endpoint para mantener Railway caliente (evita cold starts)
# ---------------------------------------------------------------------------
@api_router.get("/ping")
async def ping():
    return {"pong": True}


# ---------------------------------------------------------------------------
# Wire up
# ---------------------------------------------------------------------------
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=CORS_ORIGINS or ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    logger.info("SoundOps Pro API ready — Firestore + Cloudinary + Railway")


@app.on_event("shutdown")
async def shutdown():
    logger.info("SoundOps Pro API shutdown")
